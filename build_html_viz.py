"""
Construit une visualisation HTML interactive standalone — qualité publication
— version multi-températures.

Source : un fichier xlsx par température MD, structure identique (feuille `endo`,
18 colonnes), 3 motifs A / B / A+B.

Pour ajouter une nouvelle température, ajouter une entrée à TEMPERATURES :
    ("T15K", "tableau_filled_T15K.xlsx", "15"),
    ("T35K", "tableau_filled_T35K.xlsx", "35"),

Sortie : DM/B2/viz_DM_B2.html (ouvrable hors ligne, Plotly + Tailwind via CDN)

Encodage visuel :
  - Couleur = motif (palette Okabe-Ito, daltoniens)
  - Dash    = température (solid / dash / dot / dashdot)
  - Marker  = température (circle / square / diamond / triangle)
  - Courbe MOYENNE (toggle) = ligne épaisse longdashdot + marker star
    + bande ±écart-type semi-transparente, calculée sur les T actives
    aux points temporels communs.
"""

import json
from pathlib import Path
import openpyxl

HERE = Path(__file__).parent
OUT  = HERE / "viz_DM_B2.html"

# (label, fichier xlsx, valeur K) — extensible
TEMPERATURES = [
    ("T5K",  "tableau_filled_T5K.xlsx",     "5"),
    ("T50K", "tableau_filled_v2_T50K.xlsx", "50"),
    ("T15K", "tableau_filled_T15K.xlsx",    "15"),
    ("T35K", "tableau_filled_T35K.xlsx",    "35"),
]


def load_data_file(path: Path, T_label: str, T_kelvin: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["endo"]
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({
            "T":           T_label,
            "T_K":         T_kelvin,
            "molecule":    ws.cell(row=r, column=1).value,
            "motif":       ws.cell(row=r, column=2).value,
            "t":           float(ws.cell(row=r, column=3).value),
            "EN":          ws.cell(row=r, column=4).value,
            "GEO":         ws.cell(row=r, column=5).value,
            "HOMA":        ws.cell(row=r, column=6).value,
            "HOMA_TOTAL":  ws.cell(row=r, column=7).value,
            "LDM_OFF":     ws.cell(row=r, column=8).value,
            "LDM_DIAG":    ws.cell(row=r, column=9).value,
            "LDM_FROB":    ws.cell(row=r, column=10).value,
            "LDM_RMSD":    ws.cell(row=r, column=11).value,
            "LDM_CT_PCT":  ws.cell(row=r, column=12).value,
            "Q_FROB":      ws.cell(row=r, column=13).value,
            "Q_RMSD":      ws.cell(row=r, column=14).value,
            "Q_G":         ws.cell(row=r, column=15).value,
            "S_HOM":       ws.cell(row=r, column=16).value,
            "S_HOM_G":     ws.cell(row=r, column=17).value,
            "S_G":         ws.cell(row=r, column=18).value,
        })
    for r in rows:
        r["step"] = int(round(abs(r["t"]) / 0.2))
        r["direction"] = "TS" if r["t"] == 0 else ("reverse" if r["t"] < 0 else "forward")

    # Dédup à t=0 : tableau contient 2 lignes par motif à t=0 (TS reverse + TS forward,
    # SP regénérés indépendamment → valeurs légèrement différentes). On garde la 1re
    # (TS issu du calcul reverse) → courbe continue traversant un seul point TS.
    seen, deduped, dups = set(), [], 0
    for r in rows:
        key = (r["motif"], r["t"])
        if key in seen: dups += 1; continue
        seen.add(key); deduped.append(r)
    print(f"  {T_label}: {len(rows)} lignes -> {len(deduped)} apres dedup ({dups} doublons a t=0)")
    return deduped


def load_all_data():
    out = {}
    avail = []
    for label, fname, T_K in TEMPERATURES:
        path = HERE / fname
        if not path.exists():
            print(f"  ! {fname} introuvable, {label} ignoré")
            continue
        out[label] = load_data_file(path, label, T_K)
        avail.append((label, T_K))
    return out, avail


def build_html(data_by_T: dict, available_T: list):
    data_json = json.dumps(data_by_T, ensure_ascii=False)
    # Liste des T disponibles pour le header
    T_list_human = ", ".join(f"T = {tK} K" for _, tK in available_T)

    html = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<title>DM B2 exo — Évolution multi-températures des descripteurs d'aromaticité</title>
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
<script src="https://cdn.tailwindcss.com"></script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=STIX+Two+Text:ital,wght@0,400;0,600;1,400;1,600&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
  :root {
    --serif: 'STIX Two Text', 'STIX2Text-Regular', 'Latin Modern Roman', 'Times New Roman', Times, serif;
    --ui:    'Inter', system-ui, -apple-system, 'Segoe UI', sans-serif;
  }
  body { font-family: var(--ui); }
  .panel { background: #ffffff; border-radius: 12px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
  .chip { transition: all 0.15s ease; }
  .chip-active { box-shadow: 0 0 0 2px currentColor inset; }
  .tbl-row:nth-child(odd) { background: #f8fafc; }
  .desc-card { transition: all 0.1s ease; cursor: pointer; }
  .desc-card:hover { background: #eff6ff; }
  .desc-card-active { background: #dbeafe; border-color: #2563eb; }
  .math { font-family: var(--serif); font-style: italic; }
  .modal-bg { background: rgba(15, 23, 42, 0.55); }
  #plot text { font-family: var(--serif) !important; }
</style>
</head>
<body class="bg-slate-100 min-h-screen">
<div class="max-w-[1600px] mx-auto p-6 space-y-4">

  <!-- En-tête -->
  <header class="panel p-6">
    <h1 class="text-2xl font-bold text-slate-800" style="font-family:var(--serif);">
      Évolution multi-températures des descripteurs d'aromaticité — Diels-Alder exo (B2)
    </h1>
    <div class="text-sm text-slate-600 mt-3 grid md:grid-cols-2 gap-x-8 gap-y-1.5">
      <div><span class="font-semibold text-slate-700">Système :</span>
           Nitrobenzène + 1,3-butadiène, voie exo</div>
      <div><span class="font-semibold text-slate-700">Températures disponibles :</span>
           <span id="hdr-temps">__T_LIST__</span></div>
      <div><span class="font-semibold text-slate-700">Dynamique moléculaire :</span>
           B3LYP/6-31G(d), NVE, partant du TS</div>
      <div><span class="font-semibold text-slate-700">Pas MD :</span>
           dt = 0.2 fs, snapshot tous les 10 pas (2 fs)</div>
      <div><span class="font-semibold text-slate-700">Single-point électronique :</span>
           ωB97X-D/6-311++G(d,p)</div>
      <div><span class="font-semibold text-slate-700">Analyse aromaticité :</span>
           AroX v0.2.0 (HOMA + LDM)</div>
    </div>
    <div class="flex flex-wrap gap-3 mt-4 text-xs">
      <span class="px-2.5 py-1 rounded bg-sky-100 text-sky-800 border border-sky-200">
        Reverse : t &lt; 0 (vers réactifs)
      </span>
      <span class="px-2.5 py-1 rounded bg-amber-100 text-amber-800 border border-amber-200">
        TS : t = 0
      </span>
      <span class="px-2.5 py-1 rounded bg-rose-100 text-rose-800 border border-rose-200">
        Forward : t &gt; 0 (vers produits)
      </span>
    </div>
  </header>

  <!-- Contrôles -->
  <section class="panel p-5 space-y-4">
    <div class="grid md:grid-cols-2 gap-5">
      <div>
        <div class="text-sm font-semibold text-slate-700 mb-2">
          Températures (cliquer pour activer/désactiver)
        </div>
        <div id="temp-chips" class="flex gap-2 flex-wrap"></div>
        <div class="text-[11px] text-slate-500 mt-1.5">
          <i>Encodage : style de ligne = température (solid / dash / dot / dashdot).</i>
        </div>
      </div>
      <div>
        <div class="text-sm font-semibold text-slate-700 mb-2">
          Motifs (cliquer pour activer/désactiver)
        </div>
        <div id="motif-chips" class="flex gap-2 flex-wrap"></div>
        <div class="text-[11px] text-slate-500 mt-1.5">
          <i>Encodage : couleur = motif (palette Okabe-Ito, daltoniens).</i>
        </div>
      </div>
    </div>

    <div>
      <div class="text-sm font-semibold text-slate-700 mb-2">
        Descripteurs à afficher (un panneau empilé par descripteur)
      </div>
      <div id="descriptor-grid" class="grid grid-cols-2 md:grid-cols-4 lg:grid-cols-5 gap-2"></div>
    </div>

    <div class="flex items-center gap-4 pt-3 border-t border-slate-200 flex-wrap">
      <label class="flex items-center gap-2 text-sm">
        <input type="checkbox" id="show-markers" checked class="rounded">
        <span>Marqueurs visibles</span>
      </label>
      <label class="flex items-center gap-2 text-sm">
        <input type="checkbox" id="show-ts-line" checked class="rounded">
        <span>Ligne TS (t=0)</span>
      </label>
      <label class="flex items-center gap-2 text-sm">
        <input type="checkbox" id="show-zones" checked class="rounded">
        <span>Zones colorées rev/TS/fwd</span>
      </label>
      <label class="flex items-center gap-2 text-sm">
        <input type="checkbox" id="show-minor-grid" class="rounded">
        <span>Grille mineure</span>
      </label>
      <label class="flex items-center gap-2 text-sm border-l pl-3 border-slate-300">
        <input type="checkbox" id="show-mean" class="rounded">
        <span><b>Moyenne sur les T actives</b> (par motif)</span>
      </label>
      <label class="flex items-center gap-2 text-sm">
        <input type="checkbox" id="show-std" class="rounded">
        <span>Bande ±σ</span>
      </label>

      <div class="ml-auto flex items-center gap-2">
        <button id="btn-export-png"
                class="text-sm px-3 py-1.5 rounded bg-blue-600 text-white hover:bg-blue-700 shadow-sm">
          Export PNG (×3, 300 dpi)
        </button>
        <button id="btn-export-svg"
                class="text-sm px-3 py-1.5 rounded bg-emerald-600 text-white hover:bg-emerald-700 shadow-sm">
          Export SVG vectoriel
        </button>
        <button id="btn-export-config"
                class="text-sm px-3 py-1.5 rounded bg-slate-700 text-white hover:bg-slate-800 shadow-sm">
          Export figure publication…
        </button>
        <button id="reset-btn"
                class="text-sm px-3 py-1.5 rounded bg-slate-200 hover:bg-slate-300">
          Réinitialiser
        </button>
      </div>
    </div>
  </section>

  <!-- Graphiques -->
  <section class="panel p-3">
    <div id="plot" style="width:100%; min-height:600px;"></div>
  </section>

  <!-- Tableau filtrable -->
  <section class="panel p-5">
    <div class="flex items-center gap-4 mb-3 flex-wrap">
      <h2 class="text-lg font-semibold text-slate-700">Données brutes</h2>
      <input type="search" id="tbl-search" placeholder="Filtre texte (motif, temps, T...)"
             class="border rounded px-3 py-1 text-sm flex-1 min-w-[200px]">
      <select id="tbl-temp-filter" class="border rounded px-2 py-1 text-sm">
        <option value="">Toutes T</option>
      </select>
      <select id="tbl-motif-filter" class="border rounded px-2 py-1 text-sm">
        <option value="">Tous motifs</option>
      </select>
      <label class="text-sm flex items-center gap-2">
        <span>t min</span>
        <input type="number" id="tbl-tmin" value="-200" step="10" class="border rounded px-2 py-1 w-20 text-sm">
      </label>
      <label class="text-sm flex items-center gap-2">
        <span>t max</span>
        <input type="number" id="tbl-tmax" value="200" step="10" class="border rounded px-2 py-1 w-20 text-sm">
      </label>
      <button id="tbl-export"
              class="text-sm px-3 py-1 rounded bg-emerald-100 text-emerald-800 hover:bg-emerald-200">
        Export CSV
      </button>
    </div>
    <div class="overflow-auto max-h-[400px] border rounded">
      <table id="tbl" class="text-xs w-full">
        <thead class="bg-slate-200 sticky top-0">
          <tr id="tbl-head"></tr>
        </thead>
        <tbody id="tbl-body"></tbody>
      </table>
    </div>
    <div class="text-xs text-slate-500 mt-2"><span id="tbl-count"></span></div>
  </section>

  <footer class="text-xs text-slate-400 text-center pb-6 pt-2" style="letter-spacing: 0.4em;">
    ARIA NOROOZI
  </footer>
</div>

<!-- Modale Export Figure -->
<div id="export-modal" class="fixed inset-0 modal-bg hidden items-center justify-center z-50">
  <div class="bg-white rounded-xl shadow-2xl p-6 w-[460px] max-w-[95vw]">
    <h3 class="text-lg font-semibold text-slate-800 mb-4">Export figure publication</h3>
    <div class="space-y-3 text-sm">
      <div class="grid grid-cols-3 gap-3 items-center">
        <label>Format</label>
        <select id="exp-format" class="col-span-2 border rounded px-2 py-1">
          <option value="png">PNG raster</option>
          <option value="svg">SVG vectoriel</option>
          <option value="jpeg">JPEG raster</option>
          <option value="webp">WebP raster</option>
        </select>
      </div>
      <div class="grid grid-cols-3 gap-3 items-center">
        <label>Unité</label>
        <select id="exp-unit" class="col-span-2 border rounded px-2 py-1">
          <option value="mm">mm</option>
          <option value="in">inch</option>
          <option value="px">pixel</option>
        </select>
      </div>
      <div class="grid grid-cols-3 gap-3 items-center">
        <label>Largeur</label>
        <input id="exp-width" type="number" value="180" step="1" class="col-span-2 border rounded px-2 py-1">
      </div>
      <div class="grid grid-cols-3 gap-3 items-center">
        <label>Hauteur (par panneau)</label>
        <input id="exp-height" type="number" value="70" step="1" class="col-span-2 border rounded px-2 py-1">
      </div>
      <div class="grid grid-cols-3 gap-3 items-center">
        <label>Résolution (dpi)</label>
        <input id="exp-dpi" type="number" value="300" step="50" class="col-span-2 border rounded px-2 py-1">
      </div>
      <div class="grid grid-cols-3 gap-3 items-center">
        <label>Nom de fichier</label>
        <input id="exp-name" type="text" class="col-span-2 border rounded px-2 py-1">
      </div>
      <div class="text-xs text-slate-500 pt-1">
        Astuce : pour un journal ACS / RSC, largeur 1 col = 84 mm, 2 col = 174 mm, hauteur ~ 70 mm/panneau.
      </div>
    </div>
    <div class="flex justify-end gap-2 mt-5">
      <button id="exp-cancel" class="px-3 py-1.5 rounded bg-slate-200 hover:bg-slate-300 text-sm">
        Annuler
      </button>
      <button id="exp-go" class="px-3 py-1.5 rounded bg-blue-600 text-white hover:bg-blue-700 text-sm">
        Télécharger
      </button>
    </div>
  </div>
</div>

<script>
// =================== DONNÉES ===================
// DATA = { "T5K": [...rows], "T50K": [...rows], ... }
const DATA = __DATA_PLACEHOLDER__;

// =================== CONFIG ===================
// Couleur = motif (Okabe-Ito daltoniens)
const MOTIF_COLORS = {
  "A (nitrobenzène)":       "#0072B2",
  "B (cycle en formation)": "#D55E00",
  "A+B (cycles fusionnés)": "#009E73",
};

// Style de ligne + marker = température (extensible : 4 styles distincts dispo).
// Chips palette = Okabe-Ito thermique froid -> chaud (bleu / vert / jaune / vermillon),
// daltoniens-friendly. Les courbes elles-mêmes restent colorées par MOTIF.
const TEMP_STYLES = {
  "T5K":  { chip: "#0072B2", dash: "solid",   marker: "circle",      short: "5 K"  },
  "T15K": { chip: "#009E73", dash: "dot",     marker: "diamond",     short: "15 K" },
  "T35K": { chip: "#E69F00", dash: "dashdot", marker: "triangle-up", short: "35 K" },
  "T50K": { chip: "#D55E00", dash: "dash",    marker: "square",      short: "50 K" },
};

const DESCRIPTORS = [
  { key: "HOMA",       label: "HOMA",       axis: "<i>HOMA</i>",
    legend: "HOMA (sans unité)",                     group: "aromaticité", default: true },
  { key: "S_HOM",      label: "S-hom",      axis: "<i>S</i><sub>hom</sub>",
    legend: "S-hom (sans unité)",                    group: "aromaticité", default: true },
  { key: "S_HOM_G",    label: "S-hom(G)",   axis: "<i>S</i><sub>hom</sub>(<i>G</i>)",
    legend: "S-hom(G) (sans unité)",                 group: "aromaticité", default: false },
  { key: "EN",         label: "EN",         axis: "<i>EN</i>",
    legend: "EN — composante électronique HOMA",     group: "HOMA comp.",  default: false },
  { key: "GEO",        label: "GEO",        axis: "<i>GEO</i>",
    legend: "GEO — composante géométrique HOMA",     group: "HOMA comp.",  default: false },
  { key: "HOMA_TOTAL", label: "HOMA total", axis: "<i>HOMA</i><sub>total</sub>",
    legend: "HOMA total (sans unité)",               group: "global",      default: false },
  { key: "LDM_FROB",   label: "LDM Frob",   axis: "‖<i>LDM</i>‖<sub>F</sub>",
    legend: "LDM ‖·‖ Frobenius (full matrix)",        group: "LDM brut",    default: true },
  { key: "LDM_OFF",    label: "LDM Off",    axis: "<i>LDM</i><sub>off-diag</sub>",
    legend: "LDM off-diagonal (sans unité)",         group: "LDM brut",    default: false },
  { key: "LDM_DIAG",   label: "LDM Diag",   axis: "<i>LDM</i><sub>diag</sub>",
    legend: "LDM diagonal (sans unité)",             group: "LDM brut",    default: false },
  { key: "LDM_RMSD",   label: "LDM RMSD",   axis: "<i>LDM</i><sub>RMSD</sub>",
    legend: "LDM RMSD (sans unité)",                 group: "LDM avancé",  default: false },
  { key: "LDM_CT_PCT", label: "LDM CT %",   axis: "<i>LDM</i><sub>CT</sub> (%)",
    legend: "LDM charge-transfer (%)",               group: "LDM avancé",  default: true },
  { key: "Q_FROB",     label: "Q Frob",     axis: "<i>Q</i><sub>Frob</sub>",
    legend: "Q Frobenius (sans unité)",              group: "LDM avancé",  default: false },
  { key: "Q_RMSD",     label: "Q RMSD",     axis: "<i>Q</i><sub>RMSD</sub>",
    legend: "Q RMSD (sans unité)",                   group: "LDM avancé",  default: false },
  { key: "Q_G",        label: "Q(G)",       axis: "<i>Q</i>(<i>G</i>)",
    legend: "Q(G) (sans unité)",                     group: "LDM avancé",  default: false },
  { key: "S_G",        label: "S(G)",       axis: "<i>S</i>(<i>G</i>)",
    legend: "S(G) (sans unité)",                     group: "LDM avancé",  default: false },
];

const SERIF = "STIX Two Text, STIX2Text-Regular, Latin Modern Roman, Times New Roman, Times, serif";
const T_KEYS = Object.keys(DATA);                       // températures disponibles
const ALL_MOTIFS = Object.keys(MOTIF_COLORS);

// =================== ÉTAT ===================
const state = {
  activeTemps:       new Set(T_KEYS),
  activeMotifs:      new Set(ALL_MOTIFS),
  activeDescriptors: new Set(DESCRIPTORS.filter(d => d.default).map(d => d.key)),
  showMarkers:   true,
  showTSLine:    true,
  showZones:     true,
  showMinorGrid: false,
  showMean:      false,
  showStd:       false,
};

// =================== INIT UI ===================
function initTempChips() {
  const cont = document.getElementById("temp-chips");
  cont.innerHTML = "";
  T_KEYS.forEach(T => {
    const style = TEMP_STYLES[T] || { chip: "#64748b", short: T };
    const el = document.createElement("button");
    el.className = "chip px-3 py-1.5 rounded-full text-sm font-medium border-2";
    el.style.color = style.chip;
    el.style.borderColor = style.chip;
    el.style.background = "white";
    el.textContent = T + "  (" + style.short + ")";
    el.classList.add("chip-active");
    el.addEventListener("click", () => {
      if (state.activeTemps.has(T)) {
        state.activeTemps.delete(T);
        el.classList.remove("chip-active");
        el.style.opacity = "0.35";
      } else {
        state.activeTemps.add(T);
        el.classList.add("chip-active");
        el.style.opacity = "1";
      }
      render();
    });
    cont.appendChild(el);
  });
}

function initMotifChips() {
  const cont = document.getElementById("motif-chips");
  cont.innerHTML = "";
  Object.entries(MOTIF_COLORS).forEach(([motif, color]) => {
    const el = document.createElement("button");
    el.className = "chip px-3 py-1.5 rounded-full text-sm font-medium border-2";
    el.style.color = color;
    el.style.borderColor = color;
    el.style.background = "white";
    el.textContent = motif;
    el.classList.add("chip-active");
    el.addEventListener("click", () => {
      if (state.activeMotifs.has(motif)) {
        state.activeMotifs.delete(motif);
        el.classList.remove("chip-active");
        el.style.opacity = "0.35";
      } else {
        state.activeMotifs.add(motif);
        el.classList.add("chip-active");
        el.style.opacity = "1";
      }
      render();
    });
    cont.appendChild(el);
  });
}

function initDescriptorGrid() {
  const cont = document.getElementById("descriptor-grid");
  cont.innerHTML = "";
  DESCRIPTORS.forEach(d => {
    const el = document.createElement("div");
    el.className = "desc-card border rounded p-2 text-center select-none";
    if (d.default) el.classList.add("desc-card-active");
    el.innerHTML = `<div class="math text-base">${d.axis}</div>
                    <div class="text-[10px] text-slate-500 mt-0.5">${d.group}</div>`;
    el.addEventListener("click", () => {
      if (state.activeDescriptors.has(d.key)) {
        state.activeDescriptors.delete(d.key);
        el.classList.remove("desc-card-active");
      } else {
        state.activeDescriptors.add(d.key);
        el.classList.add("desc-card-active");
      }
      render();
    });
    cont.appendChild(el);
  });
}

function initControls() {
  const bind = (id, key) => document.getElementById(id).addEventListener("change", e => {
    state[key] = e.target.checked; render();
  });
  bind("show-markers", "showMarkers");
  bind("show-ts-line", "showTSLine");
  bind("show-zones",   "showZones");
  bind("show-minor-grid", "showMinorGrid");
  bind("show-mean",    "showMean");
  bind("show-std",     "showStd");

  document.getElementById("reset-btn").addEventListener("click", () => {
    state.activeTemps       = new Set(T_KEYS);
    state.activeMotifs      = new Set(ALL_MOTIFS);
    state.activeDescriptors = new Set(DESCRIPTORS.filter(d => d.default).map(d => d.key));
    Object.assign(state, { showMarkers: true, showTSLine: true, showZones: true,
                           showMinorGrid: false, showMean: false, showStd: false });
    initTempChips(); initMotifChips(); initDescriptorGrid();
    ["show-markers","show-ts-line","show-zones"].forEach(id =>
      document.getElementById(id).checked = true);
    ["show-minor-grid","show-mean","show-std"].forEach(id =>
      document.getElementById(id).checked = false);
    render();
  });

  // Export PNG
  document.getElementById("btn-export-png").addEventListener("click", () => {
    const n = Math.max(1, state.activeDescriptors.size);
    Plotly.downloadImage("plot", { format: "png", filename: autoFilename("png"),
      width: 1600, height: 320 * n + 120, scale: 3 });
  });
  // Export SVG
  document.getElementById("btn-export-svg").addEventListener("click", () => {
    const n = Math.max(1, state.activeDescriptors.size);
    Plotly.downloadImage("plot", { format: "svg", filename: autoFilename("svg"),
      width: 1600, height: 320 * n + 120 });
  });
  // Modale Export
  const modal = document.getElementById("export-modal");
  document.getElementById("btn-export-config").addEventListener("click", () => {
    document.getElementById("exp-name").value = autoFilename("");
    modal.classList.remove("hidden"); modal.classList.add("flex");
  });
  document.getElementById("exp-cancel").addEventListener("click", () => {
    modal.classList.add("hidden"); modal.classList.remove("flex");
  });
  document.getElementById("exp-go").addEventListener("click", () => {
    const fmt   = document.getElementById("exp-format").value;
    const unit  = document.getElementById("exp-unit").value;
    const wVal  = parseFloat(document.getElementById("exp-width").value);
    const hVal  = parseFloat(document.getElementById("exp-height").value);
    const dpi   = parseFloat(document.getElementById("exp-dpi").value);
    let   name  = (document.getElementById("exp-name").value || "figure").trim();
    if (name.toLowerCase().endsWith("." + fmt)) name = name.slice(0, -(fmt.length + 1));

    const n = Math.max(1, state.activeDescriptors.size);
    const pxPerUnit = unit === "mm" ? dpi / 25.4 : unit === "in" ? dpi : 1;
    const totalW = Math.round(wVal * pxPerUnit);
    const totalH = Math.round(hVal * n * pxPerUnit);
    const scale  = unit === "px" ? Math.max(1, dpi / 96) : 1;

    Plotly.downloadImage("plot", { format: fmt, filename: name,
      width: totalW, height: totalH, scale: scale });
    modal.classList.add("hidden"); modal.classList.remove("flex");
  });
}

function autoFilename(ext) {
  const date = new Date().toISOString().slice(0, 10);
  const Ts = [...state.activeTemps].join("-");
  const descs = [...state.activeDescriptors].slice(0, 3).join("_");
  const base = `DM_B2_exo_${date}_${Ts}_${descs || "vide"}`.replace(/[^A-Za-z0-9_\-]/g, "_");
  return ext ? base + "." + ext : base;
}

// =================== HELPERS DATA ===================
function rowsByMotif(rows) {
  const out = {};
  rows.forEach(r => { (out[r.motif] = out[r.motif] || []).push(r); });
  Object.values(out).forEach(arr => arr.sort((a, b) => a.t - b.t));
  return out;
}

// Pour la moyenne : pour chaque motif, retourne { ts: [...], mean: [...], std: [...] }
// calculé aux points temporels EXACTS partagés par TOUTES les T actives.
function computeMeanPerMotif(motif, descKey) {
  const Ts = [...state.activeTemps];
  if (Ts.length < 1) return null;
  // Map: T → (Map: t → value)
  const perT = {};
  Ts.forEach(T => {
    perT[T] = new Map();
    (DATA[T] || []).filter(r => r.motif === motif).forEach(r => {
      perT[T].set(r.t, r[descKey]);
    });
  });
  // Points communs : intersection des keys
  const refTs = [...perT[Ts[0]].keys()];
  const commonTs = refTs.filter(t => Ts.every(T => perT[T].has(t)))
                        .sort((a, b) => a - b);
  if (commonTs.length === 0) return null;
  const means = [], stds = [];
  commonTs.forEach(t => {
    const vals = Ts.map(T => perT[T].get(t));
    const m = vals.reduce((a, b) => a + b, 0) / vals.length;
    const v = vals.reduce((a, b) => a + (b - m) ** 2, 0) / vals.length;
    means.push(m); stds.push(Math.sqrt(v));
  });
  return { ts: commonTs, mean: means, std: stds };
}

// =================== TRACÉ ===================
function render() {
  const activeDesc = DESCRIPTORS.filter(d => state.activeDescriptors.has(d.key));
  if (activeDesc.length === 0 || state.activeTemps.size === 0) {
    Plotly.purge("plot");
    document.getElementById("plot").innerHTML =
      '<div class="text-center py-16 text-slate-400" style="font-family:var(--serif);font-style:italic;">'
      + 'Sélectionne au moins une température et un descripteur</div>';
    renderTable();
    return;
  }

  // Bornes temporelles globales (toutes T)
  const allRows = [].concat(...Object.values(DATA));
  const tMin = Math.min(...allRows.map(r => r.t));
  const tMax = Math.max(...allRows.map(r => r.t));

  const traces = [];

  activeDesc.forEach((desc, i) => {
    const yAxis = i === 0 ? "y" : "y" + (i + 1);

    // 1) Une trace par (T active × motif actif) : couleur=motif, dash+marker=T
    state.activeTemps.forEach(T => {
      const style  = TEMP_STYLES[T] || { dash: "solid", marker: "circle", short: T };
      const grouped = rowsByMotif(DATA[T] || []);
      state.activeMotifs.forEach(motif => {
        const arr = grouped[motif] || [];
        if (arr.length === 0) return;
        traces.push({
          x:      arr.map(r => r.t),
          y:      arr.map(r => r[desc.key]),
          type:   "scatter",
          mode:   state.showMarkers ? "lines+markers" : "lines",
          name:   `${motif} — ${style.short}`,
          legendgroup: `${motif}__${T}`,
          showlegend: i === 0,
          xaxis:  "x",
          yaxis:  yAxis,
          line:   { color: MOTIF_COLORS[motif], width: 2.2, dash: style.dash, shape: "linear" },
          marker: { color: MOTIF_COLORS[motif], size: 6.5, symbol: style.marker,
                    line: { width: 0.7, color: "#ffffff" } },
          hoverlabel: { font: { family: SERIF, size: 13 } },
          hovertemplate:
            `<b>${motif}</b>  ·  <b>${style.short}</b><br>` +
            `<i>${desc.label}</i> = %{y:.4f}<br>` +
            `t = %{x:+.1f} fs<br>` +
            `MD step = %{customdata[0]} (%{customdata[1]})` +
            `<extra></extra>`,
          customdata: arr.map(r => [r.step, r.direction]),
        });
      });
    });

    // 2) Courbes MOYENNE (par motif, sur les T actives) + bande ±σ
    if (state.showMean && state.activeTemps.size >= 1) {
      state.activeMotifs.forEach(motif => {
        const m = computeMeanPerMotif(motif, desc.key);
        if (!m) return;
        const color = MOTIF_COLORS[motif];

        // Bande ±σ (deux traces fill='tonexty') — uniquement si ≥ 2 T actives
        if (state.showStd && state.activeTemps.size >= 2) {
          const upper = m.mean.map((v, k) => v + m.std[k]);
          const lower = m.mean.map((v, k) => v - m.std[k]);
          traces.push({
            x: m.ts, y: upper,
            type: "scatter", mode: "lines",
            xaxis: "x", yaxis: yAxis,
            line: { color: color, width: 0 },
            hoverinfo: "skip", showlegend: false,
            legendgroup: `mean__${motif}`,
          });
          traces.push({
            x: m.ts, y: lower,
            type: "scatter", mode: "lines",
            xaxis: "x", yaxis: yAxis,
            line: { color: color, width: 0 },
            fill: "tonexty", fillcolor: hexToRgba(color, 0.13),
            hoverinfo: "skip", showlegend: false,
            legendgroup: `mean__${motif}`,
          });
        }

        // Courbe moyenne — ligne épaisse longdashdot + marker star
        traces.push({
          x: m.ts, y: m.mean,
          type: "scatter",
          mode: state.showMarkers ? "lines+markers" : "lines",
          name: `⟨${motif}⟩ moyenne`,
          legendgroup: `mean__${motif}`,
          showlegend: i === 0,
          xaxis: "x", yaxis: yAxis,
          line:   { color: color, width: 3.5, dash: "longdashdot" },
          marker: { color: color, size: 9, symbol: "star",
                    line: { width: 1.2, color: "#1e293b" } },
          hoverlabel: { font: { family: SERIF, size: 13 } },
          hovertemplate:
            `<b>⟨${motif}⟩</b>  moyenne sur ${state.activeTemps.size} T<br>` +
            `<i>${desc.label}</i> = %{y:.4f} ± %{customdata:.4f}<br>` +
            `t = %{x:+.1f} fs<extra></extra>`,
          customdata: m.std,
        });
      });
    }
  });

  // Layout
  const n = activeDesc.length;
  const gap = 0.055;
  const totalGap = gap * (n - 1);
  const h = (1 - totalGap) / n;
  const layout = {
    grid: { rows: n, columns: 1, pattern: "independent" },
    height: Math.max(360 + 300 * n, 540),
    margin: { l: 90, r: 40, t: 110, b: 80 },
    hovermode: "closest",
    font: { family: SERIF, size: 14, color: "#1e293b" },
    legend: {
      orientation: "h",
      x: 0.5, xanchor: "center",
      y: 1.04, yanchor: "bottom",
      font: { family: SERIF, size: 13 },
      bgcolor: "rgba(255,255,255,0)", bordercolor: "rgba(0,0,0,0)",
      itemwidth: 30,
    },
    plot_bgcolor: "#ffffff",
    paper_bgcolor: "#ffffff",
    shapes: [],
    annotations: [],
  };

  activeDesc.forEach((desc, i) => {
    const yKey   = i === 0 ? "yaxis" : "yaxis" + (i + 1);
    const yRef   = i === 0 ? "y" : "y" + (i + 1);
    const yDomTop = 1 - i * (h + gap);
    const yDomBot = yDomTop - h;
    layout[yKey] = {
      title: { text: desc.axis, font: { family: SERIF, size: 16, color: "#0f172a" }, standoff: 12 },
      domain: [yDomBot, yDomTop],
      gridcolor: "#e2e8f0", gridwidth: 1,
      zerolinecolor: "#94a3b8", zerolinewidth: 1,
      tickfont: { family: SERIF, size: 13, color: "#334155" },
      ticks: "outside", ticklen: 5, tickwidth: 1, tickcolor: "#64748b",
      showline: true, linewidth: 1, linecolor: "#334155", mirror: true,
      automargin: true,
      minor: state.showMinorGrid
        ? { showgrid: true, gridcolor: "#f1f5f9", gridwidth: 0.5, ticklen: 3, tickcolor: "#94a3b8" }
        : { showgrid: false },
    };

    if (state.showTSLine) {
      layout.shapes.push({
        type: "line", x0: 0, x1: 0, xref: "x", yref: `${yRef} domain`,
        y0: 0, y1: 1, line: { color: "#94a3b8", width: 1.5, dash: "dash" }, layer: "above",
      });
    }

    if (state.showZones) {
      layout.shapes.push(
        { type: "rect", xref: "x", yref: `${yRef} domain`,
          x0: tMin - 5, x1: -6, y0: 0, y1: 1,
          fillcolor: "#0072B2", opacity: 0.04, line: { width: 0 }, layer: "below" },
        { type: "rect", xref: "x", yref: `${yRef} domain`,
          x0: -6, x1: 6, y0: 0, y1: 1,
          fillcolor: "#E69F00", opacity: 0.10, line: { width: 0 }, layer: "below" },
        { type: "rect", xref: "x", yref: `${yRef} domain`,
          x0: 6, x1: tMax + 5, y0: 0, y1: 1,
          fillcolor: "#CC79A7", opacity: 0.04, line: { width: 0 }, layer: "below" },
      );
    }

    if (i === 0 && state.showTSLine) {
      layout.annotations.push({
        x: 0, y: 1.015, xref: "x", yref: `${yRef} domain`,
        text: "TS", showarrow: false,
        font: { color: "rgba(71,85,105,0.7)", size: 12, family: SERIF, style: "italic" },
      });
    }
  });

  layout.xaxis = {
    title: {
      text: "Temps  <i>t</i>  (fs)  —  TS à <i>t</i> = 0  —  reverse ← | → forward",
      font: { family: SERIF, size: 15, color: "#0f172a" }, standoff: 14,
    },
    range: [tMin - 5, tMax + 5],
    gridcolor: "#e2e8f0", gridwidth: 1, zeroline: false,
    tickfont: { family: SERIF, size: 13, color: "#334155" },
    ticks: "outside", ticklen: 5, tickwidth: 1, tickcolor: "#64748b",
    showline: true, linewidth: 1, linecolor: "#334155", mirror: true,
    anchor: "y" + n, automargin: true,
    minor: state.showMinorGrid
      ? { showgrid: true, gridcolor: "#f1f5f9", gridwidth: 0.5, ticklen: 3, tickcolor: "#94a3b8" }
      : { showgrid: false },
  };
  for (let i = 1; i < n; i++) {
    layout["xaxis" + (i + 1)] = {
      matches: "x", showticklabels: false, showgrid: true, gridcolor: "#e2e8f0",
      showline: true, linewidth: 1, linecolor: "#334155", mirror: true,
      anchor: "y" + (i + 1),
    };
  }

  const config = {
    displaylogo: false, responsive: true,
    toImageButtonOptions: { format: "png", filename: autoFilename(""), scale: 3 },
    modeBarButtonsToRemove: ["sendDataToCloud", "lasso2d", "select2d"],
  };

  Plotly.react("plot", traces, layout, config);
  renderTable();
}

function hexToRgba(hex, alpha) {
  const m = hex.replace("#", "");
  const n = parseInt(m, 16);
  const r = (n >> 16) & 255, g = (n >> 8) & 255, b = n & 255;
  return `rgba(${r},${g},${b},${alpha})`;
}

// =================== TABLEAU ===================
const TBL_COLS = [
  ["T", "T (K)"], ["motif", "Motif"], ["t", "t (fs)"], ["step", "Step"], ["direction", "Dir"],
  ["HOMA", "HOMA"], ["EN", "EN"], ["GEO", "GEO"], ["HOMA_TOTAL", "HOMA total"],
  ["LDM_FROB", "LDM Frob"], ["LDM_OFF", "LDM Off"], ["LDM_DIAG", "LDM Diag"],
  ["LDM_RMSD", "LDM RMSD"], ["LDM_CT_PCT", "LDM CT %"],
  ["Q_FROB", "Q Frob"], ["Q_RMSD", "Q RMSD"], ["Q_G", "Q(G)"],
  ["S_HOM", "S-hom"], ["S_HOM_G", "S-hom(G)"], ["S_G", "S(G)"],
];

function getAllRows() {
  return [].concat(...T_KEYS.map(T => (DATA[T] || []).map(r => ({ ...r }))));
}
const ALL_ROWS_FLAT = getAllRows();

function initTableHead() {
  const tr = document.getElementById("tbl-head");
  tr.innerHTML = TBL_COLS.map(c => `<th class="px-2 py-1 text-left">${c[1]}</th>`).join("");
  const selT = document.getElementById("tbl-temp-filter");
  T_KEYS.forEach(T => {
    const opt = document.createElement("option");
    opt.value = T; opt.textContent = T + " (" + (TEMP_STYLES[T]?.short || T) + ")";
    selT.appendChild(opt);
  });
  const selM = document.getElementById("tbl-motif-filter");
  Object.keys(MOTIF_COLORS).forEach(m => {
    const opt = document.createElement("option");
    opt.value = m; opt.textContent = m;
    selM.appendChild(opt);
  });
  ["tbl-search","tbl-temp-filter","tbl-motif-filter","tbl-tmin","tbl-tmax"].forEach(id =>
    document.getElementById(id).addEventListener("input", renderTable));
  document.getElementById("tbl-export").addEventListener("click", exportCSV);
}

function getFilteredRows() {
  const q = document.getElementById("tbl-search").value.toLowerCase();
  const Tf = document.getElementById("tbl-temp-filter").value;
  const motifFilter = document.getElementById("tbl-motif-filter").value;
  const tmin = parseFloat(document.getElementById("tbl-tmin").value);
  const tmax = parseFloat(document.getElementById("tbl-tmax").value);
  return ALL_ROWS_FLAT.filter(r => {
    if (Tf && r.T !== Tf) return false;
    if (motifFilter && r.motif !== motifFilter) return false;
    if (!isNaN(tmin) && r.t < tmin) return false;
    if (!isNaN(tmax) && r.t > tmax) return false;
    if (q) {
      const txt = `${r.T} ${r.T_K} ${r.motif} ${r.t} ${r.step} ${r.direction}`.toLowerCase();
      if (!txt.includes(q)) return false;
    }
    return true;
  });
}

function renderTable() {
  const rows = getFilteredRows().sort((a, b) =>
    a.T.localeCompare(b.T) || a.t - b.t || a.motif.localeCompare(b.motif));
  const body = document.getElementById("tbl-body");
  body.innerHTML = rows.map(r => {
    const tds = TBL_COLS.map(([k]) => {
      let v = r[k];
      if (k === "T") v = (TEMP_STYLES[r.T]?.short) || r.T_K;
      else if (typeof v === "number") {
        if (k === "t") v = v.toFixed(1);
        else if (k === "step") v = String(v);
        else v = v.toFixed(4);
      }
      let style = "";
      if (k === "motif") style = `style="color:${MOTIF_COLORS[r.motif]};font-weight:600"`;
      if (k === "T")     style = `style="color:${TEMP_STYLES[r.T]?.chip || '#64748b'};font-weight:600"`;
      return `<td class="px-2 py-1" ${style}>${v ?? ""}</td>`;
    }).join("");
    return `<tr class="tbl-row">${tds}</tr>`;
  }).join("");
  document.getElementById("tbl-count").textContent =
    `${rows.length} lignes affichées sur ${ALL_ROWS_FLAT.length} total`;
}

function exportCSV() {
  const rows = getFilteredRows().sort((a, b) =>
    a.T.localeCompare(b.T) || a.t - b.t || a.motif.localeCompare(b.motif));
  const sep = ",";
  const head = TBL_COLS.map(c => c[1]).join(sep);
  const body = rows.map(r => TBL_COLS.map(([k]) => {
    let v = r[k];
    if (k === "T") v = r.T_K;
    if (typeof v === "string" && v.includes(sep)) return `"${v.replace(/"/g, '""')}"`;
    return v ?? "";
  }).join(sep)).join("\n");
  const blob = new Blob([head + "\n" + body], { type: "text/csv;charset=utf-8" });
  const a = document.createElement("a");
  a.href = URL.createObjectURL(blob);
  a.download = "DM_B2_descripteurs_multiT.csv";
  a.click();
}

// =================== BOOT ===================
initTempChips();
initMotifChips();
initDescriptorGrid();
initControls();
initTableHead();
render();
</script>
</body>
</html>
"""

    html = html.replace("__DATA_PLACEHOLDER__", data_json)
    html = html.replace("__T_LIST__", T_list_human)
    return html


def main():
    print("Chargement des fichiers température…")
    data, available = load_all_data()
    total = sum(len(v) for v in data.values())
    print(f"Total : {total} lignes ({len(data)} température(s) : {', '.join(data.keys())})")
    html = build_html(data, available)
    OUT.write_text(html, encoding="utf-8")
    print(f"Visualisation HTML écrite : {OUT}")
    print(f"  taille : {OUT.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
